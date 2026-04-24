import os
import pickle
import json
import csv
import logging
from dotenv import load_dotenv
from pypdf import PdfReader
import docx2txt
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import pytesseract
    from PIL import Image
except ImportError:
    pytesseract = None
from rank_bm25 import BM25Okapi
from openai import OpenAI
from src.core.database import db
from duckduckgo_search import DDGS

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("app.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("RAG-System")

load_dotenv()

class RAGEngine:
    def __init__(self):
        self.data_dir = os.getenv("DATA_DIR", "./data")
        self.index_file = os.path.join(os.getenv("VECTOR_DB_DIR", "./chroma_db"), "bm25_index.pkl")
        self.model_name = os.getenv("DEFAULT_MODEL", "google/gemma-2-9b-it:free")
        self.api_key = os.getenv("OPENROUTER_API_KEY")
        
        self.client = OpenAI(
            base_url="https://openrouter.ai/api/v1",
            api_key=self.api_key,
            default_headers={
                "HTTP-Referer": "https://localhost:3000",
                "X-Title": "Local RAG Chatbot"
            }
        )
        
        self.chunks = [] # List of {"content": str, "source": str}
        self.bm25 = None
        self.load_index()

    def load_index(self):
        if os.path.exists(self.index_file):
            logger.info("Loading existing index...")
            try:
                with open(self.index_file, "rb") as f:
                    data = pickle.load(f)
                    self.chunks = data["chunks"]
                    self.bm25 = data["bm25"]
            except Exception as e:
                logger.error(f"Error loading index: {e}")

    def save_index(self):
        os.makedirs(os.path.dirname(self.index_file), exist_ok=True)
        with open(self.index_file, "wb") as f:
            pickle.dump({"chunks": self.chunks, "bm25": self.bm25}, f)

    def extract_text(self):
        all_docs = []
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
            
        for root, dirs, files in os.walk(self.data_dir):
            for file in files:
                file_path = os.path.join(root, file)
                text = ""
                try:
                    if file.endswith(".pdf"):
                        reader = PdfReader(file_path)
                        for page in reader.pages:
                            text += page.extract_text() + "\n"
                    elif file.endswith(".docx"):
                        text = docx2txt.process(file_path)
                    elif file.endswith((".png", ".jpg", ".jpeg")):
                        if pytesseract:
                            try:
                                text = pytesseract.image_to_string(Image.open(file_path), lang='vie+eng')
                            except Exception:
                                text = f"[Loi OCR: Khong the doc file anh {file}]"
                        else:
                            text = "[Loi: Thu vien OCR chua duoc cai dat]"
                    elif file.endswith(".xlsx"):
                        if openpyxl:
                            wb = openpyxl.load_workbook(file_path, data_only=True)
                            text = ""
                            for sheet in wb.sheetnames:
                                ws = wb[sheet]
                                for row in ws.iter_rows(values_only=True):
                                    text += "\t".join([str(c) if c is not None else "" for c in row]) + "\n"
                        else:
                            text = "[Loi: Thu vien openpyxl chua duoc cai dat]"
                    elif file.endswith(".csv"):
                        with open(file_path, "r", encoding="utf-8") as f:
                            reader = csv.reader(f)
                            text = "\n".join(["\t".join(row) for row in reader])
                    elif file.endswith((".txt", ".md")):
                        with open(file_path, "r", encoding="utf-8") as f:
                            text = f.read()
                    
                    if text.strip():
                        all_docs.append({"content": text, "source": file})
                except Exception as e:
                    logger.error(f"Error reading {file}: {e}")
        return all_docs

    def ingest_data(self):
        logger.info(f"Ingesting data from {self.data_dir}...")
        docs = self.extract_text()
        if not docs: return "No documents found."

        self.chunks = []
        for doc in docs:
            content = doc["content"]
            source = doc["source"]
            # Basic chunking
            lines = content.split("\n")
            current_chunk = ""
            for line in lines:
                if len(current_chunk) + len(line) < 1000:
                    current_chunk += line + "\n"
                else:
                    self.chunks.append({"content": current_chunk.strip(), "source": source})
                    current_chunk = line + "\n"
            if current_chunk:
                self.chunks.append({"content": current_chunk.strip(), "source": source})

        tokenized_corpus = [c["content"].lower().split() for c in self.chunks]
        self.bm25 = BM25Okapi(tokenized_corpus)
        self.save_index()
        return f"Successfully ingested {len(docs)} docs, {len(self.chunks)} chunks."

    def web_search(self, query):
        """Search DuckDuckGo and return consolidated text."""
        logger.info(f"Performing web search for: {query}")
        try:
            with DDGS() as ddgs:
                results = [r for r in ddgs.text(query, max_results=5)]
                if not results: return None
                search_text = "\n\n".join([f"Source: {r['href']}\nContent: {r['body']}" for r in results])
                return search_text
        except Exception as e:
            logger.error(f"Web search error: {e}")
            return None

    def query(self, user_input, session_id="default"):
        from dotenv import load_dotenv
        load_dotenv(override=True)
        self.model_name = os.getenv("DEFAULT_MODEL", "google/gemma-2-9b-it:free")
        custom_system_prompt = os.getenv("SYSTEM_PROMPT", "Bạn là một trợ lý AI thông minh.")
        
        # 1. Get History
        history = db.get_history(session_id, limit=6)
        history_str = "\n".join([f"{r}: {c}" for r, c in history])

        # 2. Local Retrieval
        context = ""
        sources = set()
        is_web_result = False

        if self.bm25 and self.chunks:
            tokenized_query = user_input.lower().split()
            scores = self.bm25.get_scores(tokenized_query)
            if max(scores) > 0.5: # Threshold for local confidence
                top_indices = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)[:5]
                top_chunks = [self.chunks[i] for i in top_indices if scores[i] > 0]
                
                # Kiểm tra tương thích giữa format cũ (string) và mới (dict)
                processed_chunks = []
                for c in top_chunks:
                    if isinstance(c, dict):
                        processed_chunks.append(c.get("content", ""))
                        if c.get("source"): sources.add(c["source"])
                    else:
                        processed_chunks.append(str(c))
                        sources.add("Tài liệu cũ (Vui lòng Re-index)")
                
                context = "\n---\n".join(processed_chunks)
                logger.info(f"Local retrieval successful. Sources: {sources}")
        
        # 3. Fallback to Web Search if no local results
        if not context:
            web_context = self.web_search(user_input)
            if web_context:
                context = web_context
                is_web_result = True
                logger.info("Using web search results.")

        if not context:
            return "Xin lỗi, tôi không tìm thấy thông tin này trong tài liệu cục bộ và cũng không thể tra cứu trực tuyến lúc này."

        # 4. Prepare Prompt
        source_instruction = "Trích dẫn tên file nguồn nếu thông tin đến từ Context cục bộ." if not is_web_result else "Trích dẫn link nguồn từ Internet."
        
        system_instructions = f"""{custom_system_prompt}
Dựa trên thông tin được cung cấp (Context), hãy trả lời câu hỏi của người dùng.

YÊU CẦU ĐỊNH DẠNG (BẮT BUỘC):
1. Ưu tiên sử dụng Bảng (Markdown Table) nếu có nhiều thông tin so sánh hoặc liệt kê thuộc tính.
2. Sử dụng Danh sách có dấu chấm đầu dòng (Bullet points) để trình bày ý chính rõ ràng.
3. In đậm (Bold) các từ khóa quan trọng hoặc tiêu đề.
4. Sử dụng Markdown Code block nếu có đoạn code.
5. {source_instruction}

Nếu thông tin không có trong tài liệu, hãy nói rằng bạn không biết."""

        full_prompt = f"Context:\n{context}\n\nHistory:\n{history_str}\n\nQuestion: {user_input}"

        try:
            response = self.client.chat.completions.create(
                model=self.model_name,
                messages=[
                    {"role": "system", "content": system_instructions},
                    {"role": "user", "content": full_prompt}
                ]
            )
            answer = response.choices[0].message.content

            # Append sources if available
            if sources:
                answer += f"\n\n📚 **Nguồn tài liệu:** {', '.join(sources)}"
            elif is_web_result:
                answer += "\n\n🌐 **Nguồn:** Tra cứu trực tuyến"

            db.add_message(session_id, "User", user_input)
            db.add_message(session_id, "Assistant", answer)
            return answer
        except Exception as e:
            logger.error(f"Query error: {e}")
            return f"Lỗi hệ thống: {str(e)}"

rag_engine = RAGEngine()
